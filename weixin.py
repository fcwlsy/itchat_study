import itchat
from itchat.content import *
from openpyxl import load_workbook
from openpyxl import Workbook
import time
import os


# 群聊信息监听
@itchat.msg_register([TEXT, PICTURE, RECORDING, ATTACHMENT, VIDEO], isGroupChat=True)
def information(msg):

    msg_time = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()) #收到信息的时间
    msg_room = msg['User']['NickName'] #信息来自哪个群
    msg_user = msg['ActualNickName'] #谁发的信息
    #print(msg_room,rooms)
    if msg_room in rooms:
        #获得信息的具体内容
        if msg['Type'] == 'Text': #文本信息
            msg_content = msg['Content'] #信息内容
            link = ''
        elif msg['Type'] == 'Picture' \
                or msg['Type'] == 'Recording' \
                or msg['Type'] == 'Video' \
                or msg['Type'] == 'Attachment':
            msg_content = r"" + msg['FileName'] #图片等的标题
            msg['Text'](msg_room + '/' + msg['FileName']) #下载图片等
            link = msg['FileName']
    
        mylist = [] #临时存储要保存的信息列表
        mylist.append(msg_time)
        mylist.append(msg_room)
        mylist.append(msg_user)
        mylist.append(msg_content)
        msg_file = msg_room + '/record.xlsx'
        wb1 = load_workbook(msg_file)
        ws1 = wb1['Sheet']
        rows = ws1.max_row + 1
        #print("行号:{}".format(rows))
        ws1.append(mylist)
        ws1["D%d" % rows].hyperlink = link
        wb1.save(msg_file)
    
        print(msg_time,msg_room,msg_user,msg_content)
if __name__ == '__main__':
    itchat.auto_login(hotReload=True)
    chatrooms = itchat.get_chatrooms(update=True, contactOnly=True)
    print('正在监测的群聊：', len(chatrooms), '个')
    rooms = []
    for item in chatrooms:
        print(item['NickName'])
        rooms.append(item['NickName'])
        msg_dir = item['NickName']
        
        if not os.path.exists(msg_dir):
            os.mkdir(msg_dir)
        if not os.path.exists(msg_dir + '/record.xlsx'):
            wb = Workbook()
            ws = wb.active
            ws.column_dimensions["A"].width = 22
            ws.column_dimensions["D"].width = 100
            ws.append(['接收时间','来自群组','发送者','信息内容'])
            wb.save(msg_dir + '/record.xlsx')
    itchat.run()